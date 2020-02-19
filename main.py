import datetime
import requests
from bs4 import BeautifulSoup
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
import sys
import shutil

try:
    market = sys.argv[1]
    underlying = sys.argv[2]
except Exception:
    pass

months = {
    1: 'January',
    2: 'February',
    3: 'March',
    4: 'April',
    5: 'May',
    6: 'June',
    7: 'July',
    8: 'August',
    9: 'September',
    10: 'October',
    11: 'November',
    12: 'December'
}

# market = 'fund'
# underlying = 'SPY'

print(f'Type: {market}  Underlying: {underlying}')

today = datetime.datetime.today()
filename: str = market + "_" + underlying + "_" + str(today)
filename = filename.replace(" ", "_").replace(".", "-").replace(":", "-") + ".xlsx"
print(f'Creating workbook')
stonks = Workbook()
first = True
for i in range(5):
    delta = relativedelta(months=i)
    date = today + delta
    month_name = months[date.month]
    year = date.year
    expires_index = 0

    URL = f'https://www.marketwatch.com/investing/{market}/{underlying}/OptionsMonth?month={month_name}&year={year}&countrycode=US'
    page = requests.get(URL)
    soup = BeautifulSoup(page.content, 'html.parser')
    rows = soup.find_all('tr')
    print("Parsing data")
    sheet_row = 0
    move_row = False
    for row_i, row in enumerate(rows):
        cells = row.find_all('td')
        sheet_row += 1
        for col_i, cell in enumerate(cells):
            classes = cell.get("class")
            val = cell.text.strip()
            if "CALLS" in val:
                print(f'Creating worksheet {year}-{date.month} #{expires_index}')
                if first:
                    ws = stonks.active
                    ws.title = f'{year}-{date.month} #{expires_index}'
                    first = False
                else:
                    ws = stonks.create_sheet(title=f'{year}-{date.month} #{expires_index}')
                expires_index += 1
                sheet_row = 0
            if "Expires " in val:
                ws.title = f'{year}-{date.month} {val[8:]}'
            if "Current price as" in val:
                move_row = True

            if classes is not None and not ("acenter" in classes or "aleft" in classes):
                val = val.replace(',', '')
                try:
                    val = float(val)
                except Exception:
                    pass
            ws.cell(column=col_i + 1, row=sheet_row + 1, value=val)

            if move_row:
                ws.move_range(f'A{sheet_row + 1}:O{sheet_row + 1}', rows=-sheet_row, cols=15)
                sheet_row -= 1
                move_row = False

print(f"Saving to file {filename}")
stonks.save(filename=filename)
shutil.copy(filename, market + "_" + underlying + "_latest.xlsx")
